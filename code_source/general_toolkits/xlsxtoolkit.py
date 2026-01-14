#!/usr/bin/env python3 

from typing import Optional, List
from contextlib import contextmanager
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

import time

import polars as pl

@staticmethod
@contextmanager
def open_worksheet(
    str_path_file:str=None,
    str_name_sheet:str='single_variant',
    bol_read_only:bool=False,
    bol_data_only:bool=False,
    func_print:Optional[callable]=print,
):
    """  上下文管理器：安全打开和关闭worksheet 
    Args:
        str_path_file: Excel文件路径
        str_name_sheet: 工作表名称
        bol_read_only: 是否以只读模式打开
        bol_data_only: 是否只读取数据值（不读取公式）
        func_print: 打印函数
    """
    obj_workbook = None
    obj_worksheet = None
    try:
        obj_workbook = load_workbook(
            str_path_file, 
            read_only=bol_read_only,
            data_only=bol_data_only
        )
        obj_worksheet = obj_workbook[str_name_sheet]
        yield obj_workbook, obj_worksheet
    except Exception as e:
        raise e
    finally:
        if obj_workbook is not None:
            try:
                obj_workbook.close()
            except Exception as e:
                func_print(f"关闭workbook时出错: {e}")
                raise e
    return obj_workbook, obj_worksheet

@staticmethod
def catch_coordinate_Anchor(
    str_path_file:str=None,
    str_name_sheet='single_variant',
    str_name_Anchor = "Anchor_Yearly",
    n_row_max_search:Optional[int]=500,
    n_col_max_search:Optional[int]=200,
    func_print:Optional[callable]=print,
):
    """ 读取excel文件中Anchor的坐标
    """
    tpl_coordinate = None
    with open_worksheet(
        str_path_file=str_path_file,
        str_name_sheet=str_name_sheet,
        bol_read_only=True,
        bol_data_only=True,
        func_print=func_print,
    ) as (obj_workbook, obj_worksheet):
            
        # 一次性读取所有数据（只读值，不创建Cell对象）
        lst_content_row = list(obj_worksheet.iter_rows(
            min_row=1,
            max_row=n_row_max_search,
            min_col=1,
            max_col=n_col_max_search,
            values_only=True,
        ))
        
    # 在内存中快速搜索
    for idx_row, lst_row in enumerate(lst_content_row):
        for idx_col, value in enumerate(lst_row):
            if value and str_name_Anchor in str(value):
                col_letter = get_column_letter(idx_col + 1)
                tpl_coordinate = (col_letter, idx_row + 1, value)

        # 检查是否找到锚点
    if tpl_coordinate is None:
        raise ValueError(f"未找到包含 {str_name_Anchor} 的单元格")
    return tpl_coordinate

@staticmethod
def _infer_range_col(
    obj_worksheet : Worksheet,
    idx_col_table_begin: int,
    idx_col_table_end: int = None,
    idx_row_reference: int = 0,
) -> int:
    """ 推断数据的列实际范围, 通过检查参考行是否为连续的数据: 不为空或None
    
    Args:
        obj_worksheet: openpyxl worksheet对象
        idx_col_table_begin: 列起始索引(0-based)
        idx_col_table_end: 列结束索引(0-based), None则使用worksheet最大列
        idx_row_reference: 参考行索引(0-based), 用于判断列是否有数据
    
    Returns:
        int: 实际列结束索引(0-based)
    """
    # 确定搜索范围
    n_col_max_search = idx_col_table_end if idx_col_table_end and idx_col_table_end > 0 else obj_worksheet.max_column - 1
    
    # 读取参考行的数据（一次性读取，避免多次访问）
    # 使用 iter_rows 读取单行数据
    lst_value_row = None
    for row_tuple in obj_worksheet.iter_rows(
        min_row=idx_row_reference + 1,
        max_row=idx_row_reference + 1,
        min_col=idx_col_table_begin + 1,
        max_col=n_col_max_search + 1,
        values_only=True
    ):
        # iter_rows 返回行的迭代器，每行是一个 tuple
        # 我们只读取一行，所以直接取第一个（也是唯一一个）tuple
        lst_value_row = list(row_tuple)
        break
    
    # 如果没有读取到数据，返回起始列
    if lst_value_row is None:
        raise ValueError("未能读取参考行的数据, 读取为空")
    
    # 从后向前查找最后一个非空值的位置
    idx_col_table_end_infer = idx_col_table_begin
    for idx_col_offset in range(len(lst_value_row) - 1, -1, -1):
        value = lst_value_row[idx_col_offset]
        # 判断是否为有效值（非空且非空白字符串）
        if value is not None and not (isinstance(value, str) and value.strip() == ''):
            idx_col_table_end_infer = idx_col_table_begin + idx_col_offset
            break
    
    return idx_col_table_end_infer

@staticmethod
def _infer_range_row(
    obj_worksheet : Worksheet,
    idx_row_table_begin: int,
    idx_row_table_end: int = None,
    idx_col_reference: int = 0,
) -> int:
    """ 推断数据的行实际范围, 通过检查参考列是否为连续的数据: 不为空或None

    Args:
        obj_worksheet: openpyxl worksheet对象
        idx_row_table_begin: 行起始索引(0-based)
        idx_row_table_end: 行结束索引(0-based), None则使用worksheet最大行
        idx_col_reference: 参考列索引(0-based), 用于判断行是否有数据
    
    Returns:
        int: 实际行结束索引(0-based)
    """
    # 确定搜索范围
    n_row_max_search = idx_row_table_end if idx_row_table_end and idx_row_table_end > 0 else obj_worksheet.max_row - 1
    
    # 读取参考列的数据（一次性读取，避免多次访问）
    lst_value_col = []
    for cell in obj_worksheet.iter_rows(
        min_row=idx_row_table_begin + 1,
        max_row=n_row_max_search + 1,
        min_col=idx_col_reference + 1,
        max_col=idx_col_reference + 1,
        values_only=True
    ):
        # iter_rows返回的是行的迭代器，每行是一个tuple
        # 因为只有一列，所以每个tuple只有一个元素
        lst_value_col.append(cell[0] if cell else None)
    
    # 从后向前查找最后一个非空值的位置
    idx_row_table_end_infer = idx_row_table_begin
    for idx_row_offset in range(len(lst_value_col) - 1, -1, -1):
        value = lst_value_col[idx_row_offset]
        # 判断是否为有效值（非空且非空白字符串）
        if value is not None and not (isinstance(value, str) and value.strip() == ''):
            idx_row_table_end_infer = idx_row_table_begin + idx_row_offset
            break
    
    return idx_row_table_end_infer

@staticmethod
def _infer_table_range(
    obj_worksheet:Worksheet,
    idx_row_header:int=0,
    idx_row_table_begin:int=0,
    idx_row_table_end:int=0,
    idx_col_table_begin:int=0,
    idx_col_table_end:int=0,
):
    """ 推断表格的实际范围
    """
    # 推断范围
    idx_col_table_end = _infer_range_col(
        obj_worksheet=obj_worksheet,
        idx_col_table_begin=idx_col_table_begin,
        idx_col_table_end=idx_col_table_end,
        idx_row_reference=idx_row_header,
    )
    idx_row_table_end = _infer_range_row(
        obj_worksheet=obj_worksheet,
        idx_row_table_begin=idx_row_table_begin,
        idx_row_table_end=idx_row_table_end,
        idx_col_reference=idx_col_table_begin,
    )
    return idx_row_table_begin, idx_row_table_end, idx_col_table_begin, idx_col_table_end

@staticmethod
def _read_data_range(
    obj_worksheet:Worksheet,
    idx_row_header:int=0,
    idx_row_table_begin:int=0,
    idx_row_table_end:int=0,
    idx_col_table_begin:int=0,
    idx_col_table_end:int=0,
    bol_infer_range:bool=True,
) -> pl.DataFrame:
    """ 读取指定范围的数据为Polars DataFrame
    """
    # 推断范围并读取数据
    if bol_infer_range:
        _, idx_row_table_end, _, idx_col_table_end = _infer_table_range(
            obj_worksheet=obj_worksheet,
            idx_row_header=idx_row_header,
            idx_row_table_begin=idx_row_table_begin,
            idx_row_table_end=idx_row_table_end,
            idx_col_table_begin=idx_col_table_begin,
            idx_col_table_end=idx_col_table_end,
        )

    # 读取范围的数据
    lst_data = []
    for row in obj_worksheet.iter_rows(
        min_row=idx_row_table_begin + 1,
        max_row=idx_row_table_end + 1,
        min_col=idx_col_table_begin + 1,
        max_col=idx_col_table_end + 1,
        values_only=True
    ):
        # 按行读取
        lst_data.append(list(row))
    
    # 转置列表的行列以便按列访问
    lst_data = list(zip(*lst_data))
    
    # 读取表头
    lst_header = []
    for row in obj_worksheet.iter_rows(
        min_row=idx_row_header + 1,
        max_row=idx_row_header + 1,
        min_col=idx_col_table_begin + 1,
        max_col=idx_col_table_end + 1,
        values_only=True
    ):
        lst_header += list(row)
    
    # 创建Polars DataFrame
    dic_data = {
        str_name_col : lst_data[idx_col]
        for idx_col, str_name_col in enumerate(lst_header) 
    }

    df_data = pl.DataFrame(
        data=dic_data,
        strict=False,
    )
    return df_data, (idx_row_table_begin, idx_row_table_end, idx_col_table_begin, idx_col_table_end)

@staticmethod
def get_table_range_Anchor(
    str_path_file:str=None,
    str_name_sheet:str='single_variant',
    str_name_Anchor:str="Anchor_Yearly",
    n_offset_header_row:int=5,
    n_offset_data_row_begin:int=8,
    n_offset_data_col_begin:int=1,
    n_rows_max_read:int=50,
    n_cols_max_read:int=200,
):
    """ 获取包含表格数据的锚点的范围
    """
    # 获取锚点坐标
    str_col_letter_Anchor, str_row_number_Anchor, _ = catch_coordinate_Anchor(
        str_path_file=str_path_file,
        str_name_sheet=str_name_sheet,
        str_name_Anchor=str_name_Anchor
    )
    # 转为python的行列索引
    idx_col_Anchor = column_index_from_string(str_col_letter_Anchor) - 1
    idx_row_Anchor = int(str_row_number_Anchor) - 1
    # 计算行序号
    idx_row_header = idx_row_Anchor + n_offset_header_row
    idx_row_table_begin = idx_row_Anchor + n_offset_data_row_begin
    idx_row_table_end = idx_row_table_begin + n_rows_max_read
    # 计算列序号
    idx_col_table_begin = idx_col_Anchor + n_offset_data_col_begin
    idx_col_table_end = idx_col_table_begin + n_cols_max_read

    # 读取数据
    with open_worksheet(
        str_path_file=str_path_file,
        str_name_sheet=str_name_sheet,
        bol_read_only=True,
        bol_data_only=True,
    ) as (obj_workbook, obj_worksheet):
        # 读取有效范围
        _, idx_row_table_end, _, idx_col_table_end = _infer_table_range(
            obj_worksheet=obj_worksheet,
            idx_row_header=idx_row_header,
            idx_row_table_begin=idx_row_table_begin,
            idx_row_table_end=idx_row_table_end,
            idx_col_table_begin=idx_col_table_begin,
            idx_col_table_end=idx_col_table_end,
        )
    # 计算行列数
    n_rows_table_read = idx_row_table_end - idx_row_table_begin + 1
    n_cols_table_read = idx_col_table_end - idx_col_table_begin + 1
    return idx_row_table_begin, idx_row_table_end, idx_col_table_begin, idx_col_table_end, n_rows_table_read, n_cols_table_read

@staticmethod
def read_table_Anchor(
    str_path_file:str=None,
    str_name_sheet:str='single_variant',
    str_name_Anchor:str="Anchor_Yearly",
    n_offset_header_row:int=5,
    n_offset_data_row_begin:int=8,
    n_offset_data_col_begin:int=1,
    n_rows_max_read:int=50,
    n_cols_max_read:int=200,
):
    """ 读取包含表格数据的锚点的表格数据
    """

    # 获取锚点坐标
    str_col_letter_Anchor, str_row_number_Anchor, _ = catch_coordinate_Anchor(
        str_path_file=str_path_file,
        str_name_sheet=str_name_sheet,
        str_name_Anchor=str_name_Anchor
    )

    # 转为python的行列索引
    idx_col_Anchor = column_index_from_string(str_col_letter_Anchor) - 1
    idx_row_Anchor = int(str_row_number_Anchor) - 1
    # 计算行序号
    idx_row_header = idx_row_Anchor + n_offset_header_row
    idx_row_table_begin = idx_row_Anchor + n_offset_data_row_begin
    idx_row_table_end = idx_row_table_begin + n_rows_max_read
    # 计算列序号
    idx_col_table_begin = idx_col_Anchor + n_offset_data_col_begin
    idx_col_table_end = idx_col_table_begin + n_cols_max_read

    # 读取数据
    with open_worksheet(
        str_path_file=str_path_file,
        str_name_sheet=str_name_sheet,
        bol_read_only=True,
        bol_data_only=True,
    ) as (obj_workbook, obj_worksheet):

        # 读取有效范围
        _, idx_row_table_end, _, idx_col_table_end = _infer_table_range(
            obj_worksheet=obj_worksheet,
            idx_row_header=idx_row_header,
            idx_row_table_begin=idx_row_table_begin,
            idx_row_table_end=idx_row_table_end,
            idx_col_table_begin=idx_col_table_begin,
            idx_col_table_end=idx_col_table_end,
        )

        # 读取数据
        df_data, _ = _read_data_range(
            obj_worksheet=obj_worksheet,
            idx_row_header=idx_row_header,
            idx_row_table_begin=idx_row_table_begin,
            idx_row_table_end=idx_row_table_end,
            idx_col_table_begin=idx_col_table_begin,
            idx_col_table_end=idx_col_table_end,
            bol_infer_range=False,
        )

    return df_data, (idx_row_table_begin, idx_row_table_end, idx_col_table_begin, idx_col_table_end, idx_row_header)


@staticmethod
def write_table_from_df(
    df_data: pl.DataFrame,
    str_path_file: str = None,
    str_name_sheet: str = 'single_variant',
    idx_row_header: int = None,
    idx_row_table_begin: int = None,
    idx_row_table_end: int = None,
    idx_col_table_begin: int = None,
    idx_col_table_end: int = None,
    lst_name_col_ignore: List[str] = None,
    func_print: Optional[callable] = print,
):
    """将 Polars DataFrame 高性能写入到指定 Anchor 位置的表格
    
    Args:
        df_data: 要写入的 Polars DataFrame
        str_path_file: Excel 文件路径
        str_name_sheet: 工作表名称
        idx_row_table_begin: 表格数据起始行索引 (0-based)
        idx_row_table_end: 表格数据结束行索引 (0-based)
        idx_col_table_begin: 表格数据起始列索引 (0-based)
        idx_col_table_end: 表格数据结束列索引 (0-based)
        lst_name_col_ignore: 需要跳过的列名列表
        func_print: 打印函数
    
    Returns:
        写入的实际范围 (idx_row_begin, idx_row_end, idx_col_begin, idx_col_end)
    """
    if lst_name_col_ignore is None:
        lst_name_col_ignore = []
    
    # 7. 使用 open_worksheet 上下文管理器
    with open_worksheet(
        str_path_file=str_path_file,
        str_name_sheet=str_name_sheet,
        bol_read_only=False,  # 写入模式
        bol_data_only=False,
        func_print=func_print,
    ) as (obj_workbook, obj_worksheet):
        # 1.读取表头
        lst_header = []
        for row in obj_worksheet.iter_rows(
            min_row=idx_row_header + 1,
            max_row=idx_row_header + 1,
            min_col=idx_col_table_begin + 1,
            max_col=idx_col_table_end + 1,
            values_only=True
        ):
            lst_header += list(row)
        print(f"读取表头: {lst_header}")

        # 2.验证表头和df_data列名一致(包括顺序)
        for idx_col, str_name_col in enumerate(df_data.columns):
            str_name_header = lst_header[idx_col]
            if str_name_header != str_name_col:
                raise ValueError(f"表头列名不匹配, 位置{idx_col}: df中列名 '{str_name_col}', 实际excel表头 '{str_name_header}'")
            
        # 1. 预处理:确定哪些列需要写入
        lst_bool_write_col = [
            str_name_col not in lst_name_col_ignore
            for str_name_col in df_data.columns
        ]
        
        # 9. 批量写入数据（使用 iter_rows 高性能写入）
        # 获取写入区域的单元格迭代器
        iter_range_cell = obj_worksheet.iter_rows(
            min_row=idx_row_table_begin + 1,
            max_row=idx_row_table_end + 1,
            min_col=idx_col_table_begin + 1,
            max_col=idx_col_table_end + 1,
        )
        
        # 逐行赋值（比逐单元格快很多）
        for row_excel, row_data in zip(iter_range_cell, df_data.rows()):
            # 逐行
            for cell, value, bol_write in zip(row_excel, row_data, lst_bool_write_col):
                # 一行中的逐列
                if bol_write:  # ← 只写入允许的列
                    cell.value = value
        
        # 10. 保存文件
        obj_workbook.save(str_path_file)
    return
