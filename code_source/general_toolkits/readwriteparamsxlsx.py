#!/usr/bin/python3
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
# from openpyxl.utils.cell import coordinate_from_string
# from openpyxl.utils import column_index_from_string
# from openpyxl.utils.cell import coordinate_from_string
from typing import List, Dict, Any, Tuple
import os
import re
import json
import time
# import pandas as pd
import numpy as np
import copy

class ReadWriteParamsXLSX:
    """ 读取或写入Excel文件中的参数，返回嵌套字典结构
    适用于参数配置文件的读取
    """
    def __init__(self,
        bol_time_benchmark : bool=False,
        dic_RegExp_skip_sheet : Dict[str, str]=None,
    ):
        """
        """
        self.str_path_current = os.getcwd()
        # 设定excel表格的构造格式
        self.dic_map_idx_col_to_name = {
            1 : 'param_key',
            2 : 'unit',
            3 : 'value',
            4 : 'desc'
        }
        # 设定读取表格时需要跳过的行
        self.n_row_skip_read = 2
        # 设定列序号作为参数名
        self.idx_col_key_param = 0
        # 是否启用时间基准测试
        self.bol_time_benchmark = bol_time_benchmark
        # 需要跳过的sheet名称的正则表达式字典
        if dic_RegExp_skip_sheet is None:
            self.dic_RegExp_skip_sheet = {
                'skip_begin_with_Sheet' : r'^Sheet.*',
                'skip_边际成本' : r'^边际成本.*',
            }
        pass

    @staticmethod
    def _print_time_usage(
        str_text: str="",
        flt_time_last : float=0.0,
        bol_time_benchmark : bool = True,
    ):
        """ 打印时间使用情况
        """
        if not bol_time_benchmark:
            return flt_time_last
        flt_time_current = time.time()
        flt_time_used = flt_time_current - flt_time_last
        print(str_text.format(flt_time_used))
        return flt_time_current

    @staticmethod
    def _find_max_effective_row_in_sheet(
        arr_content_sheet: np.ndarray,
        n_row_skip_read: int,
        idx_row_max_sheet: int,
    ):
        """ 找到sheet中有效数据的最大行号
        Args
        ----
            arr_content_sheet: sheet内容的numpy数组表示
            n_row_skip_read: 读取页时跳过的行数
            idx_row_max_sheet: sheet的最大行号
        """
        ## 读取idx_col_key_param所在列的实际数据范围, 并筛选出有效行
        arr_key_row = arr_content_sheet[:, 0]
        arr_idx_none = np.where(
            np.logical_or(arr_key_row == 'None', arr_key_row is None))[0]

        if arr_idx_none.size > 0:
            idx_first_none = int(arr_idx_none[0])
            idx_row_effective_max = idx_first_none + n_row_skip_read
        else:
            idx_row_effective_max = idx_row_max_sheet
        return idx_row_effective_max
    
    @staticmethod
    def _cleanup_param_keys(
        arr_content_sheet : np.array,
        idx_col_key_param : int = 0
    ) -> str:
        """ 清理参数key列进行清理
        1. 去除首尾空格和不可见字符
        """
        # 1.去除首尾空格
        arr_content_sheet[:, idx_col_key_param] = np.char.strip(
            arr_content_sheet[:, idx_col_key_param].astype(str)
        )
        return arr_content_sheet

    @staticmethod
    def _check_skip_sheet(
        str_name_sheet: str,
        dic_RegExp_skip_sheet: Dict[str, str],
    ) -> bool:
        """ 检查是否需要跳过该sheet
        """
        bol_skip = False
        for str_name_item, str_RegExp in dic_RegExp_skip_sheet.items():
            if re.match(str_RegExp, str_name_sheet):
                bol_skip = True
                break
        return bol_skip

    @staticmethod
    def _save_array_to_dict(
        arr_content_sheet : np.array,
        str_sheet_name : str,
        n_row_skip_read : int,
        idx_col_key_param : int,
        dic_map_idx_col_to_name : Dict[int, str],
    ):
        """ 保存array内容到嵌套字典结构

        Args
        ---
            arr_content_sheet: numpy数组表示的sheet内容
            dic_map_idx_col_to_name: 映射列序号到列名, 列序号从遵照openpyxl从1开始计数
            存储中包含:
                sheet_name: sheet名称
                n_row_skip_read: 读取页时跳过的行数
                idx_col_key_param : 作为param名字的key的列索引
                n_cols: 数据总列数
                n_rows: 数据总行数
                map_idx_col_to_name: 列序号到列名的映射字典
                map_param_to_np_index: 变量名到np行号的映射字典
        """
        dic_data_json_sheet = {}
        n_rows, n_cols = arr_content_sheet.shape
        # 加入info
        dic_data_json_sheet['sheet_name'] = str_sheet_name
        dic_data_json_sheet['n_row_skip_read'] = n_row_skip_read
        dic_data_json_sheet['idx_col_key_param'] = idx_col_key_param
        dic_data_json_sheet['n_cols'] = n_cols
        dic_data_json_sheet['n_rows'] = n_rows
        # 加入列名映射字典
        dic_data_json_sheet['map_idx_col_to_name'] = copy.deepcopy(dic_map_idx_col_to_name)
        # 加入变量名到np号的映射字典
        arr_param_keys = arr_content_sheet[:, idx_col_key_param]
        dic_data_json_sheet['map_param_to_np_index'] = {
            str(str_param_key) : int(idx_row)
            for idx_row, str_param_key in enumerate(arr_param_keys)
        }
        # 加入每一列的数据
        dic_data_json_sheet['data'] = {}
        for idx_col, idx_col_original in enumerate(dic_map_idx_col_to_name.keys()):
            str_name_col = dic_map_idx_col_to_name[idx_col_original]
            dic_data_json_sheet['data'][str_name_col] = arr_content_sheet[:, idx_col].tolist()
        return dic_data_json_sheet
    
    @staticmethod
    def _read_param_from_single_sheet_openpyxl_range(
        ws : Worksheet,
        str_path_file_excel_param: str,
        str_name_sheet: str,
        n_row_skip_read: int = 0,
        idx_col_key_param: int = 0,
        dic_map_idx_col_to_name: Dict[int, str] = None,
        bol_record_position: bool = True,
        bol_time_benchmark: bool = False,
    ) -> Dict[str, Any]:
        """ 从单个excel sheet中读取参数, 返回嵌套参数字典（包含位置信息）
        Args:
            str_path_file_excel_param: Excel文件路径
            str_name_sheet: sheet名称
            n_row_skip_read: 读取页时跳过的行数
            idx_col_key_param: 作为param名字的key的列索引
            dic_map_idx_col_to_name: 映射列序号到列名, 同时字典key名用于筛选需要的列
            str_name_index_df: 索引列名称
            bol_record_position: 是否记录单元格位置信息
        
        Returns:
            dict: 嵌套字典结构
                - 第一层key: 该sheet第一列的值（参数名）
                - 第二层包含:
                    - 'data': dict，包含该行的所有数据（列名为key）
                    - 'position': dict，包含每个字段的位置信息 {'列名': {'row': 行号, 'col': 列号, 'coord': 单元格坐标}}
        """
        # 初始化
        
        # 使用 with 语句打开工作簿，确保自动关闭
        flt_TimeStamp_start = time.time()

        flt_TimeStamp_last = ReadWriteParamsXLSX._print_time_usage(
            "打开工作表 {} 用时: {{}} 秒".format(str_name_sheet),
            flt_TimeStamp_start,
            bol_time_benchmark
        )
        
        # 获取需要处理的列序号
        lst_idx_col = sorted(dic_map_idx_col_to_name.keys())
        idx_col_min = min(lst_idx_col)
        idx_col_max = max(lst_idx_col)

        # 获取sheet的最大行列范围
        idx_row_min = n_row_skip_read + 1
        idx_row_max_sheet = ws.max_row
        ## 读取sheet的数据, 存为一个list
        arr_content_sheet = np.array(
            list(
                ws.iter_rows(
                    min_row=idx_row_min,
                    max_row=idx_row_max_sheet,
                    min_col=idx_col_min,
                    max_col=idx_col_max,
                    values_only=True)
            ),
            dtype=str)
        # 只保留需要的列
        lst_idx_col_in_np = [
            idx_col - min(lst_idx_col)
            for idx_col in lst_idx_col
        ]
        arr_content_sheet = arr_content_sheet[:, lst_idx_col_in_np]
        ## 读取idx_col_key_param所在列的实际数据范围, 并筛选出有效行
        idx_row_effective_max = ReadWriteParamsXLSX._find_max_effective_row_in_sheet(
            arr_content_sheet=arr_content_sheet,
            n_row_skip_read=n_row_skip_read,
            idx_row_max_sheet=idx_row_max_sheet
        )

        # 截取有效行
        arr_content_sheet = arr_content_sheet[0: idx_row_effective_max - n_row_skip_read, :]

        # 处理有效行名
        arr_content_sheet = ReadWriteParamsXLSX._cleanup_param_keys(
            arr_content_sheet=arr_content_sheet,
            idx_col_key_param=idx_col_key_param
        )

        # 转为json文件
        dic_data_json_sheet = ReadWriteParamsXLSX._save_array_to_dict(
            arr_content_sheet=arr_content_sheet,
            str_sheet_name=str_name_sheet,
            n_row_skip_read=n_row_skip_read,
            idx_col_key_param=idx_col_key_param,
            dic_map_idx_col_to_name=dic_map_idx_col_to_name
        )

        # benchmark
        flt_TimeStamp_last = ReadWriteParamsXLSX._print_time_usage(
            "读取工作表 {} 用时: {{}} 秒".format(str_name_sheet),
            flt_TimeStamp_last,
            bol_time_benchmark
        )
        
        return dic_data_json_sheet

    @staticmethod
    def _read_param_from_xlsx(
        str_path_file_excel_param: str,
        n_row_skip_read: int = 0,
        idx_col_key_param: int = 0,
        dic_RegExp_skip_sheet: Dict[str, str] = None,
        dic_map_idx_col_to_name: Dict[int, str] = None,
        bol_time_benchmark: bool = None,
    ):
        """ 读取Excel文件中的参数, 返回嵌套字典结构
        适用于参数配置文件的读取

        Args:
        -----
            str_path_file_excel_param: Excel文件路径
            n_row_skip_read: 读取页时跳过的行数
            idx_col_key_param: 变量名所在列的列序号
            dic_RegExp_skip_sheet: 需要跳过的sheet名称的正则表达式字典
            dic_map_idx_col_to_name: 映射列序号到列名, 同时字典key名用于筛选需要的列
            bol_time_benchmark: 是否启用时间基准测试
        """
        flt_time_start = time.time()
        # 初始化
        wb = None
        dic_data_json_all_sheets = {}
        # print(type(arr_content_all_sheets))
        try:
            # 打开工作簿
            wb = load_workbook(str_path_file_excel_param, data_only=True, read_only=True)
            # 遍历所有sheet
            for str_name_sheet in wb.sheetnames:
                # 检查是否需要跳过该sheet
                bol_skip = ReadWriteParamsXLSX._check_skip_sheet(
                    str_name_sheet=str_name_sheet,
                    dic_RegExp_skip_sheet=dic_RegExp_skip_sheet,
                )
                if bol_skip:
                    continue
                # 读取当前sheet
                ws = wb[str_name_sheet]
                # 读取单个sheet的参数
                dic_data_json_single_sheet = ReadWriteParamsXLSX._read_param_from_single_sheet_openpyxl_range(
                    ws = ws,
                    str_path_file_excel_param = str_path_file_excel_param,
                    str_name_sheet = str_name_sheet,
                    n_row_skip_read = n_row_skip_read,
                    idx_col_key_param = idx_col_key_param,
                    dic_map_idx_col_to_name = dic_map_idx_col_to_name,
                    bol_record_position = True,
                    bol_time_benchmark=bol_time_benchmark,
                )
                # print('dtype of arr_content_single_sheet:', type(arr_content_single_sheet))
                # 合并到总的参数sheet中
                dic_data_json_all_sheets[str_name_sheet] = dic_data_json_single_sheet
        finally:
            if wb is not None:
                wb.close()
        flt_time_last = ReadWriteParamsXLSX._print_time_usage(
            "读取Excel文件 {}\n 用时: {{}} 秒".format(str_path_file_excel_param),
            flt_time_start, bol_time_benchmark
        )
        return dic_data_json_all_sheets

    @staticmethod
    def _is_bool_string(
        str_val: str
    ) -> bool:
        """ 判断一个字符串是否表示布尔值
        """
        set_true = {"true"}
        set_false = {"false"}
        return str_val.lower() in set_true | set_false

    @staticmethod
    def _recognize_dtype_data(
        val_data : str
    ):
        """ 识别数据的类型
        把字符串转换为对应的类型
        """
        try:
            if ReadWriteParamsXLSX._is_bool_string(val_data):
                # 优先识别布尔类型
                val_converted = val_data
            else:
                val_converted = float(val_data)
                if val_converted.is_integer():
                    val_converted = int(val_converted)
            return val_converted
        except (ValueError, TypeError):
            val_converted = str(val_data).strip()
        return val_converted

    @staticmethod
    def _get_content(
        dic_all_sheet : Dict[str, Any],
        str_name_sheet : str,
        str_param_key : str,
        str_key_domain : str = 'data',
        str_content_to_get : str = None,
        bol_type_convert : bool = True,
        bol_time_benchmark : bool = False
    ) -> Any:
        """ 获取json数据的内容
        如果 str_content_to_get 为 None，则返回整个数据字典

        Args
        ----
            dic_all_sheet: 嵌套字典结构的参数, 即json文件
            str_name_sheet: 页签名称
            str_param_key: 参数名称
            str_key_domain: 数据域名称，默认为 'data'
            str_content_to_get: 需要获取的内容字段名称
            bol_type_convert: 是否尝试进行类型转换，默认为 True

        """
        # benchmark
        flt_time_start = time.time()
        # 读取指定页签的数据字典
        dic_sheet = dic_all_sheet[str_name_sheet]
        # 读取变量名映射np行序号的字典
        dic_map_param_to_np_index = dic_sheet['map_param_to_np_index']
        idx_row_param = dic_map_param_to_np_index[str_param_key]
        # 读取对应行的数据字典
        dic_data_domain = dic_sheet[str_key_domain]
        flt_time_last = ReadWriteParamsXLSX._print_time_usage(
            "获取参数 {}.{} 用时: {{}} 秒".format(str_name_sheet, str_param_key),
            flt_time_start, bol_time_benchmark
        )

        # 读取值
        if str_content_to_get is not None:
            val = dic_data_domain[str_content_to_get][idx_row_param]
            val_converted = val
            # 尝试类型转换
            if bol_type_convert:
                val_converted = ReadWriteParamsXLSX._recognize_dtype_data(val)
            result = val_converted
        else:
            result = dic_data_domain
        # benchmark
        flt_time_last = ReadWriteParamsXLSX._print_time_usage(
            "读取值用时: {} 秒",
            flt_time_last, bol_time_benchmark)
        return result

    @staticmethod
    def _get_value(
        dic_all_sheet : Dict[str, Any],
        str_name_sheet : str,
        str_param_key : str,
        str_key_domain : str = 'data',
        str_content_to_get : str = 'value',
        bol_type_convert : bool = True,
        bol_time_benchmark : bool = False
    ) -> Any:
        """ 获取数据值
        """
        return ReadWriteParamsXLSX._get_content(
            dic_all_sheet = dic_all_sheet,
            str_name_sheet = str_name_sheet,
            str_param_key = str_param_key,
            str_key_domain = str_key_domain,
            str_content_to_get = str_content_to_get,
            bol_type_convert = bol_type_convert,
            bol_time_benchmark = bol_time_benchmark
        )
    
    @staticmethod
    def _get_position_in_xlsx(
        dic_all_sheet : Dict[str, Any],
        str_name_sheet : str,
        str_param_key : str,
        str_content_to_get : str = None,
        bol_time_benchmark : bool = False
    ) -> Any:
        """ 获取所要数据值在json文件中的位置
        n_row_skip_read: 读取页时跳过的行数
        idx_col_key_param : 作为param名字的key的列索引
        n_cols: 数据总列数
        n_rows: 数据总行数
        """
        # benchmark
        flt_time_start = time.time()
        # 读取指定页签的数据字典
        dic_sheet = dic_all_sheet[str_name_sheet]
        # 获取读表需要跳过的行
        n_row_skip_read = dic_sheet['n_row_skip_read']
        # 读取数据映射
        dic_map_param_to_np_index = dic_sheet['map_param_to_np_index']
        idx_np_index = dic_map_param_to_np_index[str_param_key]
        # 计算在Excel中的行号
        idx_row_param_in_xlsx = idx_np_index + n_row_skip_read + 1
        # 计算str_content_to_get所在的列号
        dic_map_idx_col_to_name = dic_sheet['map_idx_col_to_name']
        dic_map_name_to_idx_col = {
            str_name_col : idx_col
            for idx_col, str_name_col in dic_map_idx_col_to_name.items()
        }
        idx_col_content_in_xlsx = dic_map_name_to_idx_col[str_content_to_get]
        # 转化为Excel单元格坐标
        str_letter_col = get_column_letter(idx_col_content_in_xlsx)
        str_coord_in_xlsx = f"{str_letter_col}{idx_row_param_in_xlsx}"
        # Benchmark
        flt_time_last = ReadWriteParamsXLSX._print_time_usage(
            "获取参数位置 {}.{} 用时: {{}} 秒".format(str_name_sheet, str_param_key),
            flt_time_start, bol_time_benchmark
        )
        return idx_row_param_in_xlsx, idx_col_content_in_xlsx, str_coord_in_xlsx
    
    @staticmethod
    def _get_len_params(
        dic_all_sheet : Dict[str, Any],
    ):
        """ 获取dic_all_sheet中参数总数
        """
        n_params = 0
        for str_sheet_name, dic_sheet in dic_all_sheet.items():
            n_params += dic_sheet['n_rows']
        return n_params

    @staticmethod
    def _export_dict_to_json(
        dic_all_sheet : Dict[str, Any],
        str_path_file_json : str
    ):
        """ 导出字典到json文件
        """
        with open(str_path_file_json, 'w', encoding='utf-8') as f_json:
            json.dump(dic_all_sheet, f_json, ensure_ascii=False, indent=4)
        pass

    @staticmethod
    def _convert_value_fit_for_heeds(
        val_param : Any,
        str_replace_newline : str = '__NEWLINE__',
        str_replace_comma : str = '__COMMA__'
    ):
        """ 转换参数值以适应HEEDS的要求
        """
        # 去掉值中的\n和逗号，防止CSV格式错误
        if isinstance(val_param, str):
            val_param = val_param.replace('\n', str_replace_newline).replace(',', str_replace_comma)
        elif isinstance(val_param, bool):
            val_param = str(val_param).upper()
        return val_param
    
    @staticmethod
    def _convert_value_back_from_heeds(
        val_param : Any,
        str_replace_newline : str = '__NEWLINE__',
        str_replace_comma : str = '__COMMA__'
    ):
        """ 将HEEDS格式的参数值转换回原始格式
        """
        if isinstance(val_param, str):
            val_param = val_param.replace(str_replace_newline, '\n').replace(str_replace_comma, ',')
        else:
            val_param = str(val_param)
        return val_param

    @staticmethod
    def _convert_dic_param_to_array_for_heeds(
        dic_all_sheet : Dict[str, Any],
    ) -> np.ndarray:
        """ 转换参数字典为HEEDS可识别的array格式
        """
        n_params = ReadWriteParamsXLSX._get_len_params(dic_all_sheet)
        # 初始化array
        arr_params = np.empty((n_params, 3), dtype=object)
        idx_row_current = 0
        # 循环填充array
        for str_name_sheet, dic_single_sheet in dic_all_sheet.items():
            dic_data_sheet = dic_single_sheet['data']
            lst_name_param = dic_data_sheet['param_key']
            lst_value_param = dic_data_sheet['value']
            n_params_sheet = len(lst_name_param)
            # 转换参数值以适应HEEDS的要求
            lst_value_param = [
                ReadWriteParamsXLSX._convert_value_fit_for_heeds(val_param)
                for val_param in lst_value_param
            ]
            # 填充数据
            arr_params[idx_row_current: idx_row_current + n_params_sheet, 0] = str_name_sheet
            arr_params[idx_row_current: idx_row_current + n_params_sheet, 1] = lst_name_param
            arr_params[idx_row_current: idx_row_current + n_params_sheet, 2] = lst_value_param
            # 更新当前行号
            idx_row_current += n_params_sheet
        return arr_params

    @staticmethod
    def _export_csv_file_for_heeds(
        dic_all_sheet : Dict[str, Any],
        str_path_file_csv : str,
        bol_time_benchmark : bool = False,
    ):
        """ 导出参数为HEEDS可识别的CSV文件格式
        """
        flt_time_start = time.time()
        # 转换为HEEDS格式的array
        arr_params = ReadWriteParamsXLSX._convert_dic_param_to_array_for_heeds(
            dic_all_sheet=dic_all_sheet
        )

        # 写入csv文件
        np.savetxt(
            str_path_file_csv, 
            arr_params, 
            delimiter=',', 
            fmt='%s',
            header='Sheet_Name,Parameter_Name,Value',
            comments='',  # 防止header前加'#'
            encoding='utf-8'
        )
        # Benchmark
        if bol_time_benchmark:
            flt_time_last = ReadWriteParamsXLSX._print_time_usage(
                "导出HEEDS格式CSV文件 {} 用时: {{}} 秒".format(str_path_file_csv),
                flt_time_start, bol_time_benchmark
            )
        pass

    @staticmethod
    def _save_value_array_for_heeds_to_dic_param(
        arr_params : np.ndarray,
        dic_all_sheet : Dict[str, Any],
    ) -> Dict[str, Any]:
        """ 将HEEDS格式的array中的value保存入嵌套字典结构

        Args
        ----
            arr_params: HEEDS格式的参数array
            dic_all_sheet: 此基础字典基础上更新读入array内容
        """
        # 获取变量数量和sheet名称列表
        n_params = arr_params.shape[0]
        lst_name_sheet = list(np.unique(arr_params[:, 0]))
        # 循环填充参数值
        for idx_row in range(n_params):
            str_name_sheet = arr_params[idx_row, 0]
            str_name_param = arr_params[idx_row, 1]
            val_param = arr_params[idx_row, 2]
            # 转换回原始格式
            val_param_converted = ReadWriteParamsXLSX._convert_value_back_from_heeds(val_param)
            # 更新到字典中
            dic_single_sheet = dic_all_sheet[str_name_sheet]
            dic_map_param_to_np_index = dic_single_sheet['map_param_to_np_index']
            idx_np_index = dic_map_param_to_np_index[str_name_param]
            dic_data_sheet = dic_single_sheet['data']
            # 保存值
            dic_data_sheet['value'][idx_np_index] = val_param_converted
        return dic_all_sheet

    @staticmethod
    def _save_dic_param_to_xlsx(
        dic_all_sheet : Dict[str, Any],
        str_path_file_excel_param: str,
        bol_time_benchmark: bool = False,
    ):
        """ 将嵌套字典结构的参数保存回Excel文件中

        Args
        ----
            dic_all_sheet: 嵌套字典结构的参数
            str_path_file_excel_param: Excel文件路径
            n_row_skip_read: 读取页时跳过的行数
        """
        flt_time_start = time.time()
        # 打开工作簿
        wb = load_workbook(str_path_file_excel_param)
        # 遍历所有dic_all_sheet中的sheet
        for str_name_sheet, dic_single_sheet in dic_all_sheet.items():
            # 读取工作表
            ws = wb[str_name_sheet]
            # 读取信息
            n_row_skip_read = int(dic_single_sheet["n_row_skip_read"])
            # idx_col_key_param = dic_single_sheet["idx_col_key_param"]
            # n_cols = dic_single_sheet["n_cols"]
            # n_rows = dic_single_sheet["n_rows"]
            dic_map_idx_col_to_name = dic_single_sheet["map_idx_col_to_name"]
            dic_map_name_to_idx_col = {
                str_name_col : idx_col
                for idx_col, str_name_col in dic_map_idx_col_to_name.items()
            }
            idx_value_col = int(dic_map_name_to_idx_col['value'])
            # 读取数据
            dic_data_sheet = dic_single_sheet['data']
            lst_value_param = dic_data_sheet['value']
            # 构建单元格范围并批量赋值
            idx_row_start = n_row_skip_read + 1
            idx_row_end = idx_row_start + len(lst_value_param) - 1
            
            # 一次性获取所有需要修改的单元格
            lst_cells = ws.iter_rows(
                min_row=idx_row_start,
                max_row=idx_row_end,
                min_col=idx_value_col,
                max_col=idx_value_col
            )
            lst_cells = [row for row in lst_cells]  # 转换为列表以便多次使用
            # 批量赋值
            for idx, row in enumerate(lst_cells):
                row[0].value = lst_value_param[idx]
             
        # 保存工作簿到文件
        wb.save(str_path_file_excel_param)
        wb.close()

        if bol_time_benchmark:
            flt_time_last = ReadWriteParamsXLSX._print_time_usage(
                "保存Excel文件 {} 用时: {{}} 秒".format(str_path_file_excel_param),
                flt_time_start, bol_time_benchmark
            )
        return

    @staticmethod
    def _save_csv_file_heeds_to_xlsx(
        str_path_file_csv : str,
        str_path_file_dic_param_json : str,
        str_path_file_xlsx : str,
        bol_time_benchmark : bool = False,
    ):
        """ 从HEEDS格式的CSV文件中导入参数值并保存回Excel文件中
        """
        flt_time_start = time.time()
        # 读取csv文件
        arr_params = np.loadtxt(
            str_path_file_csv,
            delimiter=',',
            dtype=str,
            skiprows=1
        )
        # 读取json文件
        with open(str_path_file_dic_param_json, 'r', encoding='utf-8') as f_json:
            dic_all_sheet = json.load(f_json)
        # 将csv文件中的值保存回字典结构
        dic_all_sheet_updated = ReadWriteParamsXLSX._save_value_array_for_heeds_to_dic_param(
            arr_params=arr_params,
            dic_all_sheet=dic_all_sheet
        )
        # 保存回Excel文件
        ReadWriteParamsXLSX._save_dic_param_to_xlsx(
            dic_all_sheet=dic_all_sheet_updated,
            str_path_file_excel_param=str_path_file_xlsx
        )
        # Benchmark
        if bol_time_benchmark:
            flt_time_last = ReadWriteParamsXLSX._print_time_usage(
                "从HEEDS格式CSV文件 {} 保存回Excel文件 {} 用时: {{}} 秒".format(
                    str_path_file_csv,
                    str_path_file_xlsx
                ),
                flt_time_start, bol_time_benchmark
            )
        return

    def read_from_xlsx(self,
        str_path_file_excel_param: str,
    ):
        """ 读取Excel文件中的参数, 返回嵌套字典结构
        适用于参数配置文件的读取

        Args:
        -----
            str_path_file_excel_param: Excel文件路径
        """
        # 时间基准测试
        bol_time_benchmark = self.bol_time_benchmark
        # 读取参数
        n_row_skip_read = self.n_row_skip_read
        idx_col_key_param = self.idx_col_key_param
        dic_RegExp_skip_sheet = copy.deepcopy(self.dic_RegExp_skip_sheet)
        dic_map_idx_col_to_name = copy.deepcopy(self.dic_map_idx_col_to_name)

        # 获取读入参数字典
        dic_data_json_all_sheets = ReadWriteParamsXLSX._read_param_from_xlsx(
            str_path_file_excel_param=str_path_file_excel_param,
            n_row_skip_read=n_row_skip_read,
            idx_col_key_param=idx_col_key_param,
            dic_RegExp_skip_sheet=dic_RegExp_skip_sheet,
            dic_map_idx_col_to_name=dic_map_idx_col_to_name,
            bol_time_benchmark=bol_time_benchmark,
        )
        return dic_data_json_all_sheets
    
    def generate_json_csv_from_xlsx(self,
        str_path_file_excel_param: str,
        str_path_file_json: str,
        str_path_file_csv: str,
    ):
        """ 从Excel文件生成json文件 和 HEEDS格式的CSV文件
        """
        bol_time_benchmark = self.bol_time_benchmark
        # 获取读入参数字典
        dic_data_json_all_sheets = self.read_from_xlsx(
            str_path_file_excel_param=str_path_file_excel_param
        )
        # 导出json文件
        ReadWriteParamsXLSX._export_dict_to_json(
            dic_all_sheet=dic_data_json_all_sheets,
            str_path_file_json=str_path_file_json
        )
        # 导出HEEDS格式的CSV文件
        ReadWriteParamsXLSX._export_csv_file_for_heeds(
            dic_all_sheet=dic_data_json_all_sheets,
            str_path_file_csv=str_path_file_csv,
            bol_time_benchmark=bol_time_benchmark
        )
        return

    def save_heeds_csv_to_xlsx(self,
        str_path_file_csv : str,
        str_path_file_dic_param_json : str,
        str_path_file_xlsx : str,
    ):
        """ 从HEEDS格式的CSV文件中导入参数值并保存回Excel文件中
        """
        bol_time_benchmark = self.bol_time_benchmark
        ReadWriteParamsXLSX._save_csv_file_heeds_to_xlsx(
            str_path_file_csv = str_path_file_csv,
            str_path_file_dic_param_json = str_path_file_dic_param_json,
            str_path_file_xlsx = str_path_file_xlsx,
            bol_time_benchmark = bol_time_benchmark
        )
        return